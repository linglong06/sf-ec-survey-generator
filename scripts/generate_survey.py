#!/usr/bin/env python3
"""
SF EC Survey Generator - SAP SuccessFactors Employee Central 调研问卷生成器

生成符合标准模板格式的Excel调研问卷，支持中英文双语。
支持从客户制度文档中提取的答案填充问卷。
支持来源追溯，记录答案来自哪个文件的哪个章节。

工作流程：
1. Claude从客户提供的制度文档中提取与SF EC相关的答案
2. 答案通过JSON文件传递给脚本（包含答案和来源信息）
3. 脚本生成带初步答案和来源的问卷Excel文件
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List, Tuple, Union

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# 默认问题库 - EC核心模块
DEFAULT_QUESTIONS = {
    "基本信息": [
        {
            "sub_category": "公司概况",
            "questions": [
                ("请简要介绍公司的业务范围和主营业务？", "Please briefly introduce the company's business scope and main business?"),
                ("公司目前的员工总数是多少？", "What is the total number of employees in the company?"),
                ("公司的组织架构层级有多少级？", "How many levels does the company's organizational structure have?"),
                ("公司是否有多个法人实体？如有，请列出。", "Does the company have multiple legal entities? If yes, please list them."),
                ("公司是否有海外分支机构或员工？", "Does the company have overseas branches or employees?"),
            ]
        },
        {
            "sub_category": "现有系统",
            "questions": [
                ("目前使用的人事管理系统是什么？", "What is the current HR management system?"),
                ("现有系统的主要痛点有哪些？", "What are the main pain points of the current system?"),
                ("是否有其他需要与SF集成的系统？", "Are there other systems that need to integrate with SF?"),
                ("现有人事数据的准确性如何？", "How accurate is the current HR data?"),
                ("数据迁移的范围和截止时间要求？", "What is the scope and deadline for data migration?"),
            ]
        },
    ],
    "组织管理": [
        {
            "sub_category": "组织架构",
            "questions": [
                ("公司的组织架构类型是什么？（事业部制/职能制/矩阵制等）", "What is the company's organizational structure type? (Business unit/Functional/Matrix, etc.)"),
                ("组织架构的调整频率如何？", "How often is the organizational structure adjusted?"),
                ("是否需要支持矩阵式汇报关系？", "Do you need to support matrix reporting relationships?"),
                ("组织单元是否有有效期管理需求？", "Do organizational units require effective date management?"),
                ("组织单元的自定义字段需求有哪些？", "What are the custom field requirements for organizational units?"),
            ]
        },
        {
            "sub_category": "组织层级",
            "questions": [
                ("公司的组织层级如何定义？（如：公司-事业部-部门-小组）", "How are organizational levels defined? (e.g., Company-BU-Department-Team)"),
                ("每个层级的命名规则是什么？", "What are the naming conventions for each level?"),
                ("组织单元的编码规则是什么？", "What are the coding rules for organizational units?"),
                ("是否有虚拟组织的需求？", "Is there a need for virtual organizations?"),
            ]
        },
        {
            "sub_category": "组织审批",
            "questions": [
                ("组织架构变更的审批流程是怎样的？", "What is the approval process for organizational changes?"),
                ("新增组织单元需要哪些必填信息？", "What are the required fields for creating a new organizational unit?"),
                ("组织合并/拆分的业务场景有哪些？", "What are the business scenarios for organization merge/split?"),
            ]
        },
    ],
    "组织信息": [
        {
            "sub_category": "组织基础信息",
            "questions": [
                ("组织单元需要维护哪些基础信息？", "What basic information needs to be maintained for organizational units?"),
                ("组织的地点/办公地址如何管理？", "How are organization locations/office addresses managed?"),
                ("组织负责人如何定义和关联？", "How are organization heads defined and linked?"),
                ("组织成本中心如何关联？", "How are organization cost centers linked?"),
            ]
        },
        {
            "sub_category": "组织属性",
            "questions": [
                ("是否需要区分不同类型的组织单元？（如：职能部门、业务部门、项目组）", "Do you need to distinguish different types of organizational units?"),
                ("组织的状态如何管理？（如：正常、冻结、关闭）", "How is organization status managed? (e.g., Active, Frozen, Closed)"),
                ("组织的预算信息是否需要在系统中管理？", "Does organization budget information need to be managed in the system?"),
                ("是否有组织层级权限控制的需求？", "Is there a need for organization-level permission control?"),
            ]
        },
    ],
    "岗位管理": [
        {
            "sub_category": "岗位体系",
            "questions": [
                ("公司目前的岗位管理体系是怎样的？", "What is the company's current position management system?"),
                ("岗位和职位(Job)的区别和管理方式是什么？", "What is the difference and management approach between Position and Job?"),
                ("岗位序列/职级体系是怎样的？", "What is the job family/grade structure?"),
                ("是否需要岗位编制管理？", "Is position headcount management required?"),
                ("岗位编码规则是什么？", "What are the position coding rules?"),
            ]
        },
        {
            "sub_category": "岗位信息",
            "questions": [
                ("岗位需要维护哪些属性信息？", "What attribute information needs to be maintained for positions?"),
                ("岗位与组织的关系是怎样的？一人一岗还是可以兼岗？", "What is the relationship between position and organization? One person one position or concurrent positions allowed?"),
                ("岗位的汇报关系如何定义？", "How is position reporting relationship defined?"),
                ("岗位是否有有效期管理？", "Is there effective date management for positions?"),
                ("岗位的自定义字段需求有哪些？", "What are the custom field requirements for positions?"),
            ]
        },
        {
            "sub_category": "岗位编制",
            "questions": [
                ("是否需要管理岗位编制数？", "Is it necessary to manage position headcount?"),
                ("编制超编时是否需要控制或预警？", "Is control or warning needed when headcount exceeds the limit?"),
                ("编制的审批流程是怎样的？", "What is the approval process for headcount changes?"),
                ("编制数据是否需要与招聘模块联动？", "Does headcount data need to integrate with the Recruiting module?"),
            ]
        },
    ],
    "人事管理": [
        {
            "sub_category": "员工基础信息",
            "questions": [
                ("员工信息需要维护哪些基础字段？", "What basic fields need to be maintained for employee information?"),
                ("是否有自定义的员工信息字段需求？", "Are there custom employee information field requirements?"),
                ("员工信息的敏感字段有哪些？如何控制访问权限？", "What are the sensitive employee fields? How is access control managed?"),
                ("员工照片的管理要求是什么？", "What are the requirements for employee photo management?"),
                ("是否需要管理员工的紧急联系人信息？", "Is it necessary to manage employee emergency contact information?"),
            ]
        },
        {
            "sub_category": "入职管理",
            "questions": [
                ("新员工入职的业务流程是怎样的？", "What is the business process for new employee onboarding?"),
                ("入职前是否需要预入职流程？", "Is a pre-onboarding process required?"),
                ("入职时需要采集哪些信息？", "What information needs to be collected during onboarding?"),
                ("入职审批流程是怎样的？", "What is the onboarding approval workflow?"),
                ("是否需要与招聘模块集成？", "Is integration with the Recruiting module required?"),
            ]
        },
        {
            "sub_category": "转正管理",
            "questions": [
                ("试用期转正的业务规则是什么？", "What are the business rules for probation confirmation?"),
                ("转正评估流程是怎样的？", "What is the probation evaluation process?"),
                ("转正审批流程是怎样的？", "What is the probation confirmation approval workflow?"),
                ("是否有试用期延长的情况？如何处理？", "Are there cases of probation extension? How are they handled?"),
            ]
        },
        {
            "sub_category": "调动管理",
            "questions": [
                ("员工调动的类型有哪些？（内部调动、跨公司调动、晋升、降职等）", "What are the types of employee transfers? (Internal transfer, Cross-company transfer, Promotion, Demotion, etc.)"),
                ("调动审批流程是怎样的？", "What is the transfer approval workflow?"),
                ("调动生效日期如何确定？", "How is the transfer effective date determined?"),
                ("调动是否需要同步更新薪酬信息？", "Does transfer need to synchronously update compensation information?"),
                ("跨公司调动涉及的工龄计算规则是什么？", "What are the seniority calculation rules for cross-company transfers?"),
            ]
        },
        {
            "sub_category": "离职管理",
            "questions": [
                ("离职类型有哪些？（主动离职、被动离职、退休等）", "What are the types of termination? (Voluntary resignation, Involuntary termination, Retirement, etc.)"),
                ("离职申请和审批流程是怎样的？", "What is the resignation application and approval workflow?"),
                ("离职交接流程是怎样的？", "What is the handover process for resignation?"),
                ("离职后员工数据的处理规则是什么？", "What are the data handling rules after employee termination?"),
                ("是否有离职面谈的需求？", "Is there a requirement for exit interviews?"),
            ]
        },
        {
            "sub_category": "合同管理",
            "questions": [
                ("劳动合同类型有哪些？", "What are the types of employment contracts?"),
                ("合同签订和续签的业务流程是怎样的？", "What is the business process for contract signing and renewal?"),
                ("合同到期提醒规则是什么？", "What are the contract expiration reminder rules?"),
                ("是否需要管理电子合同？", "Is electronic contract management required?"),
                ("合同变更流程是怎样的？", "What is the contract change process?"),
            ]
        },
        {
            "sub_category": "薪酬基础",
            "questions": [
                ("薪酬结构是怎样的？包含哪些薪酬项目？", "What is the compensation structure? What pay components are included?"),
                ("薪酬数据是否需要在EC中管理？", "Does compensation data need to be managed in EC?"),
                ("薪资调整流程是怎样的？", "What is the salary adjustment process?"),
                ("是否有多个薪资组的区分？", "Are there multiple pay groups?"),
                ("薪酬数据的保密级别如何？", "What is the confidentiality level of compensation data?"),
            ]
        },
        {
            "sub_category": "工作时间与考勤",
            "questions": [
                ("公司的工时制度是什么？（标准工时/综合工时/不定时工时）", "What is the company's working hour system? (Standard/Comprehensive/Flexible)"),
                ("是否需要在EC中管理工作时间信息？", "Is it necessary to manage working time information in EC?"),
                ("是否需要与考勤系统集成？", "Is integration with the time attendance system required?"),
                ("加班规则和管理要求是什么？", "What are the overtime rules and management requirements?"),
            ]
        },
        {
            "sub_category": "数据权限",
            "questions": [
                ("HR数据的访问权限规则是什么？", "What are the access permission rules for HR data?"),
                ("是否有按组织/字段/角色的权限控制需求？", "Is there a need for permission control by organization/field/role?"),
                ("敏感数据的脱敏规则是什么？", "What are the data masking rules for sensitive data?"),
                ("数据审批流程的权限如何控制？", "How are permissions controlled for data approval workflows?"),
            ]
        },
        {
            "sub_category": "报表与分析",
            "questions": [
                ("常用的HR报表有哪些？", "What are the commonly used HR reports?"),
                ("是否有自定义报表的需求？", "Is there a need for custom reports?"),
                ("报表数据的更新频率要求是什么？", "What are the data refresh frequency requirements for reports?"),
                ("是否需要移动端查看报表？", "Is mobile access for reports required?"),
            ]
        },
    ],
}

# 英文类别名称映射
CATEGORY_EN = {
    "基本信息": "Basic Information",
    "组织管理": "Organization Management",
    "组织信息": "Organization Information",
    "岗位管理": "Position Management",
    "人事管理": "Personnel Management",
}


def load_answers_from_json(json_path: str) -> Dict[str, Dict[str, str]]:
    """
    从JSON文件加载答案（包含答案内容和来源信息）

    JSON格式示例（新格式，支持来源）:
    {
        "请简要介绍公司的业务范围和主营业务？": {
            "answer": "公司主要从事XXX业务...",
            "source": "员工手册.pdf / 第一章 公司概况"
        },
        "公司目前的员工总数是多少？": {
            "answer": "约5000人",
            "source": "人力资源制度.docx / 第二章 组织架构"
        },
        ...
    }

    JSON格式示例（旧格式，仅答案字符串）:
    {
        "请简要介绍公司的业务范围和主营业务？": "公司主要从事XXX业务...",
        "公司目前的员工总数是多少？": "约5000人",
        ...
    }

    Args:
        json_path: JSON文件路径

    Returns:
        问题到答案和来源的映射字典 {"question": {"answer": "...", "source": "..."}}
    """
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # 转换旧格式到新格式
        result = {}
        for key, value in data.items():
            if key.startswith("_"):
                continue  # 跳过说明字段
            if isinstance(value, dict):
                # 新格式：{"answer": "...", "source": "..."}
                result[key] = {
                    "answer": value.get("answer", ""),
                    "source": value.get("source", "")
                }
            else:
                # 旧格式：直接是字符串答案
                result[key] = {
                    "answer": str(value) if value else "",
                    "source": ""
                }
        return result
    except Exception as e:
        print(f"Warning: Failed to load answers from {json_path}: {e}")
        return {}


def create_survey_excel(
    output_path: str,
    categories: Optional[list] = None,
    project_name: str = "",
    client_name: str = "",
    language: str = "both",
    include_sub_category: bool = True,
    answers: Optional[Dict[str, Dict[str, str]]] = None,
    include_answer_column: bool = True,
    include_source_column: bool = True,
) -> str:
    """
    创建SF EC调研问卷Excel文件

    Args:
        output_path: 输出文件路径
        categories: 要包含的类别列表，None表示全部
        project_name: 项目名称
        client_name: 客户名称
        language: 语言选择 - "cn"(中文), "en"(英文), "both"(双语)
        include_sub_category: 是否包含子类别行
        answers: 问题答案字典 {问题: {"answer": 答案, "source": 来源}}
        include_answer_column: 是否包含答案列
        include_source_column: 是否包含来源列

    Returns:
        生成的文件路径
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工中心 Employee Center"

    # 样式定义
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    category_font = Font(name="微软雅黑", size=11, bold=True)
    category_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    category_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    sub_category_font = Font(name="微软雅黑", size=10, bold=True, italic=True)
    sub_category_fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")

    normal_font = Font(name="微软雅黑", size=10)
    normal_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    answer_font = Font(name="微软雅黑", size=10, color="0066CC")
    answer_fill = PatternFill(start_color="F2F8FF", end_color="F2F8FF", fill_type="solid")

    source_font = Font(name="微软雅黑", size=9, color="666666", italic=True)
    source_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )

    # 确定列数和列宽
    # 列：A=序号, B=类别, C=问题中文, D=问题英文, E=答案, F=来源, G=回答
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 50

    if include_answer_column and include_source_column:
        ws.column_dimensions["E"].width = 45
        ws.column_dimensions["F"].width = 30
        ws.column_dimensions["G"].width = 30
    elif include_answer_column:
        ws.column_dimensions["E"].width = 45
        ws.column_dimensions["F"].width = 35
    else:
        ws.column_dimensions["E"].width = 35

    # 写入标题行
    if include_answer_column and include_source_column:
        headers = [
            "序号\nNumber",
            "问题类别\nQuestion Category",
            "问题名称（中文）\nQuestion Name (CN)",
            "问题名称（英文）\nQuestion Name (EN)",
            "初步答案（来自制度文档）\nPreliminary Answer",
            "来源（文件/章节）\nSource (File/Section)",
            "回答\nReply"
        ]
    elif include_answer_column:
        headers = [
            "序号\nNumber",
            "问题类别\nQuestion Category",
            "问题名称（中文）\nQuestion Name (CN)",
            "问题名称（英文）\nQuestion Name (EN)",
            "初步答案（来自制度文档）\nPreliminary Answer",
            "回答\nReply"
        ]
    else:
        headers = [
            "序号\nNumber",
            "问题类别\nQuestion Category",
            "问题名称（中文）\nQuestion Name (CN)",
            "问题名称（英文）\nQuestion Name (EN)",
            "回答\nReply"
        ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.row_dimensions[1].height = 45

    # 确定要包含的类别
    if categories is None:
        categories = list(DEFAULT_QUESTIONS.keys())

    # 写入问题
    row = 2
    question_num = 1

    # 计算最大列数
    if include_answer_column and include_source_column:
        max_col = 7
    elif include_answer_column:
        max_col = 6
    else:
        max_col = 5

    for category in categories:
        if category not in DEFAULT_QUESTIONS:
            continue

        sub_categories = DEFAULT_QUESTIONS[category]

        # 写入主类别行
        cell = ws.cell(row=row, column=1, value=question_num)
        cell.font = category_font
        cell.fill = category_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

        category_text = f"{category}"
        if language == "both":
            category_text = f"{category}\n{CATEGORY_EN.get(category, '')}"

        cell = ws.cell(row=row, column=2, value=category_text)
        cell.font = category_font
        cell.fill = category_fill
        cell.alignment = category_alignment
        cell.border = thin_border

        for col in range(3, max_col + 1):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = category_fill
            cell.border = thin_border

        ws.row_dimensions[row].height = 30
        row += 1
        question_num += 1

        # 写入子类别和问题
        for sub_cat_data in sub_categories:
            sub_category = sub_cat_data["sub_category"]
            questions = sub_cat_data["questions"]

            if include_sub_category:
                # 写入子类别行
                cell = ws.cell(row=row, column=1, value="")
                cell.fill = sub_category_fill
                cell.border = thin_border

                cell = ws.cell(row=row, column=2, value=f"  └ {sub_category}")
                cell.font = sub_category_font
                cell.fill = sub_category_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

                for col in range(3, max_col + 1):
                    cell = ws.cell(row=row, column=col, value="")
                    cell.fill = sub_category_fill
                    cell.border = thin_border

                ws.row_dimensions[row].height = 25
                row += 1

            # 写入问题
            for q_cn, q_en in questions:
                cell = ws.cell(row=row, column=1, value=question_num)
                cell.font = normal_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                cell = ws.cell(row=row, column=2, value="")
                cell.border = thin_border

                # 根据语言选择写入内容
                if language == "cn":
                    ws.cell(row=row, column=3, value=q_cn).border = thin_border
                    ws.cell(row=row, column=4, value="").border = thin_border
                elif language == "en":
                    ws.cell(row=row, column=3, value="").border = thin_border
                    ws.cell(row=row, column=4, value=q_en).border = thin_border
                else:  # both
                    ws.cell(row=row, column=3, value=q_cn).border = thin_border
                    ws.cell(row=row, column=4, value=q_en).border = thin_border

                ws.cell(row=row, column=3).font = normal_font
                ws.cell(row=row, column=3).alignment = normal_alignment
                ws.cell(row=row, column=4).font = normal_font
                ws.cell(row=row, column=4).alignment = normal_alignment

                if include_answer_column:
                    # 获取答案和来源
                    answer = ""
                    source = ""
                    if answers:
                        answer_data = answers.get(q_cn, {})
                        if isinstance(answer_data, dict):
                            answer = answer_data.get("answer", "")
                            source = answer_data.get("source", "")
                        else:
                            # 兼容旧格式（字符串）
                            answer = str(answer_data) if answer_data else ""

                    # 写入答案列
                    cell = ws.cell(row=row, column=5, value=answer)
                    cell.font = answer_font
                    if answer:
                        cell.fill = answer_fill
                    cell.alignment = normal_alignment
                    cell.border = thin_border

                    if include_source_column:
                        # 写入来源列
                        cell = ws.cell(row=row, column=6, value=source)
                        cell.font = source_font
                        if source:
                            cell.fill = source_fill
                        cell.alignment = normal_alignment
                        cell.border = thin_border

                        # 回答列
                        ws.cell(row=row, column=7, value="").border = thin_border
                    else:
                        # 回答列
                        ws.cell(row=row, column=6, value="").border = thin_border
                else:
                    # 回答列
                    ws.cell(row=row, column=5, value="").border = thin_border

                # 根据答案长度调整行高
                has_answer = answers and answers.get(q_cn)
                if has_answer:
                    ws.row_dimensions[row].height = 50
                else:
                    ws.row_dimensions[row].height = 40

                row += 1
                question_num += 1

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存文件
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_file))

    return str(output_file)


def create_answers_template(output_path: str) -> str:
    """
    创建答案JSON模板文件（新格式，包含来源字段）

    Args:
        output_path: 输出文件路径

    Returns:
        生成的文件路径
    """
    template = {
        "_说明": "这是答案模板文件。Claude会从客户制度文档中提取答案并填充此文件。",
        "_格式": "问题中文内容: {answer: 答案内容, source: 来源信息}",
        "_来源格式": "文件名 / 章节/页码，例如：员工手册.pdf / 第一章 公司概况",
    }

    # 添加所有问题
    for category, sub_cats in DEFAULT_QUESTIONS.items():
        for sub_cat_data in sub_cats:
            for q_cn, q_en in sub_cat_data["questions"]:
                template[q_cn] = {
                    "answer": "",
                    "source": ""
                }

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(template, f, ensure_ascii=False, indent=2)

    return str(output_file)


def main():
    parser = argparse.ArgumentParser(
        description="SF EC Survey Generator - 生成SAP SuccessFactors EC调研问卷",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例用法:
  # 生成基础问卷
  python generate_survey.py -o survey.xlsx

  # 带答案生成问卷（包含来源追溯）
  python generate_survey.py -o survey.xlsx -a answers.json

  # 生成答案模板
  python generate_survey.py --template answers_template.json

答案JSON格式（支持来源追溯）:
  {
    "问题？": {
      "answer": "答案内容",
      "source": "文件名.pdf / 章节"
    }
  }
        """
    )
    parser.add_argument(
        "-o", "--output",
        default="./sf_ec_survey.xlsx",
        help="输出文件路径 (默认: ./sf_ec_survey.xlsx)"
    )
    parser.add_argument(
        "-c", "--categories",
        nargs="+",
        choices=["基本信息", "组织管理", "组织信息", "岗位管理", "人事管理"],
        help="要包含的类别（默认：全部）"
    )
    parser.add_argument(
        "-p", "--project",
        default="",
        help="项目名称"
    )
    parser.add_argument(
        "--client",
        default="",
        help="客户名称"
    )
    parser.add_argument(
        "-l", "--language",
        choices=["cn", "en", "both"],
        default="both",
        help="语言选择: cn(中文), en(英文), both(双语，默认)"
    )
    parser.add_argument(
        "--no-sub-category",
        action="store_true",
        help="不包含子类别行"
    )
    parser.add_argument(
        "--list-categories",
        action="store_true",
        help="列出所有可用类别"
    )
    parser.add_argument(
        "-a", "--answers",
        help="答案JSON文件路径（由Claude从制度文档中提取）"
    )
    parser.add_argument(
        "--no-answer-column",
        action="store_true",
        help="不包含答案列"
    )
    parser.add_argument(
        "--no-source-column",
        action="store_true",
        help="不包含来源列"
    )
    parser.add_argument(
        "--template",
        help="生成答案JSON模板文件"
    )

    args = parser.parse_args()

    if args.list_categories:
        print("可用类别:")
        for cat in DEFAULT_QUESTIONS.keys():
            print(f"  - {cat}")
        return

    if args.template:
        print(f"正在生成答案模板...")
        template_file = create_answers_template(args.template)
        print(f"✅ 模板生成完成: {template_file}")
        print(f"\n使用方法:")
        print(f"  1. Claude从客户制度文档中提取答案")
        print(f"  2. 将答案填入JSON文件（包含answer和source字段）")
        print(f"  3. 运行: python generate_survey.py -o survey.xlsx -a {args.template}")
        return

    print(f"正在生成调研问卷...")
    print(f"  - 输出路径: {args.output}")
    print(f"  - 类别: {args.categories or '全部'}")
    print(f"  - 语言: {args.language}")
    print(f"  - 包含子类别: {not args.no_sub_category}")
    print(f"  - 包含答案列: {not args.no_answer_column}")
    print(f"  - 包含来源列: {not args.no_source_column}")

    # 加载答案
    answers = None
    if args.answers:
        answers = load_answers_from_json(args.answers)
        print(f"  - 加载答案: {len(answers)} 条")
        answered_count = sum(1 for v in answers.values() if isinstance(v, dict) and v.get("answer"))
        sourced_count = sum(1 for v in answers.values() if isinstance(v, dict) and v.get("source"))
        print(f"  - 有效答案: {answered_count} 条")
        print(f"  - 有来源: {sourced_count} 条")

    output_file = create_survey_excel(
        output_path=args.output,
        categories=args.categories,
        project_name=args.project,
        client_name=args.client,
        language=args.language,
        include_sub_category=not args.no_sub_category,
        answers=answers,
        include_answer_column=not args.no_answer_column,
        include_source_column=not args.no_source_column,
    )

    print(f"\n✅ 问卷生成完成: {output_file}")


if __name__ == "__main__":
    main()
